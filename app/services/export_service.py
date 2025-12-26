"""Export Service - Export reports to Excel/PDF"""
import io
import re
from datetime import date, datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """Service for exporting reports to various formats"""
    
    @staticmethod
    def _sanitize_sheet_name(name: str, max_length: int = 31) -> str:
        """
        Sanitize sheet name to be valid for Excel
        
        Excel sheet name rules:
        - Cannot contain: / \ ? * [ ] :
        - Maximum 31 characters
        - Cannot be empty
        """
        # Replace invalid characters with dash
        invalid_chars = r'[/\\?*\[\]:]'
        sanitized = re.sub(invalid_chars, '-', name)
        
        # Trim to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        
        # Ensure not empty
        if not sanitized.strip():
            sanitized = "Sheet1"
        
        return sanitized
    
    @staticmethod
    def export_to_excel(report_data: dict, report_type: str):
        """Export report data to Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title (sanitize sheet name to remove invalid characters)
        if report_type == 'daily':
            # Format date for display: convert YYYY-MM-DD to DD/MM/YYYY
            date_iso = report_data.get('date', '')
            date_display = date_iso
            if date_iso:
                try:
                    date_obj = datetime.strptime(date_iso, '%Y-%m-%d').date()
                    date_display = date_obj.strftime('%d/%m/%Y')
                except:
                    pass
            
            # Sheet title: keep ISO format (YYYY-MM-DD) which doesn't have invalid chars
            sheet_title = f"Báo cáo ngày {date_iso}"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            title = f"BÁO CÁO CHẤM CÔNG NGÀY {date_display}"
        elif report_type == 'weekly':
            sheet_title = "Báo cáo tuần"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            start_date = report_data.get('start_date', '')
            end_date = report_data.get('end_date', '')
            title = f"BÁO CÁO CHẤM CÔNG TUẦN {start_date} - {end_date}"
        else:
            month = report_data.get('month', '')
            year = report_data.get('year', '')
            # Sheet title: use dash instead of slash
            sheet_title = f"Báo cáo tháng {month}-{year}"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            # Title in cell: can use slash for display
            title = f"BÁO CÁO CHẤM CÔNG THÁNG {month}/{year}"
        
        # Title style
        title_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
        title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        
        # Header style
        header_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Write summary
        row = 3
        ws[f'A{row}'] = 'TỔNG HỢP'
        ws[f'A{row}'].font = Font(name="Arial", size=12, bold=True)
        row += 1
        
        if report_type == 'daily':
            summary_data = [
                ['Tổng nhân viên', report_data['total_employees']],
                ['Đã chấm công', report_data['checked_in']],
                ['Có mặt', report_data['present']],
                ['Đi muộn', report_data['late']],
                ['Vắng mặt', report_data['absent']],
                ['Tỷ lệ chấm công', f"{report_data['attendance_rate']}%"],
                ['Tổng giờ làm', f"{report_data['total_working_hours']}h"],
            ]
        elif report_type == 'weekly':
            summary_data = [
                ['Tổng có mặt', report_data['week_total_present']],
                ['Tổng đi muộn', report_data['week_total_late']],
                ['Tổng vắng mặt', report_data['week_total_absent']],
                ['Tổng giờ làm', f"{report_data['week_total_hours']}h"],
                ['Tỷ lệ trung bình', f"{report_data['average_attendance_rate']}%"],
            ]
        else:
            summary_data = [
                ['Tổng có mặt', report_data['total_present']],
                ['Tổng đi muộn', report_data['total_late']],
                ['Tổng vắng mặt', report_data['total_absent']],
                ['Tổng giờ làm', f"{report_data['total_working_hours']}h"],
                ['Tỷ lệ chấm công', f"{report_data['average_attendance_rate']}%"],
                ['Tỷ lệ đúng giờ', f"{report_data['on_time_rate']}%"],
            ]
        
        for item in summary_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            row += 1
        
        row += 1
        
        # Write table header
        if report_type == 'daily':
            headers = ['Nhân viên', 'Mã NV', 'Check-in', 'Check-out', 'Giờ làm', 'Trạng thái']
            ws[f'A{row}'] = headers[0]
            ws[f'B{row}'] = headers[1]
            ws[f'C{row}'] = headers[2]
            ws[f'D{row}'] = headers[3]
            ws[f'E{row}'] = headers[4]
            ws[f'F{row}'] = headers[5]
        elif report_type == 'weekly':
            headers = ['Ngày', 'Có mặt', 'Đi muộn', 'Vắng mặt', 'Giờ làm']
            ws[f'A{row}'] = headers[0]
            ws[f'B{row}'] = headers[1]
            ws[f'C{row}'] = headers[2]
            ws[f'D{row}'] = headers[3]
            ws[f'E{row}'] = headers[4]
        else:
            headers = ['Ngày', 'Có mặt', 'Đi muộn', 'Vắng mặt', 'Rời sớm']
            ws[f'A{row}'] = headers[0]
            ws[f'B{row}'] = headers[1]
            ws[f'C{row}'] = headers[2]
            ws[f'D{row}'] = headers[3]
            ws[f'E{row}'] = headers[4]
        
        # Apply header style
        for col in range(1, len(headers) + 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Write data
        if report_type == 'daily':
            for att in report_data.get('attendances', [])[:100]:  # Limit to 100 rows
                ws[f'A{row}'] = att.get('employee_name', '')
                ws[f'B{row}'] = att.get('employee_code', '')
                check_in = att.get('check_in_time', '')
                if check_in:
                    try:
                        dt = datetime.fromisoformat(check_in.replace('Z', '+00:00'))
                        ws[f'C{row}'] = dt.strftime('%H:%M')
                    except:
                        ws[f'C{row}'] = check_in[:5] if len(check_in) > 5 else check_in
                else:
                    ws[f'C{row}'] = 'N/A'
                
                check_out = att.get('check_out_time', '')
                if check_out:
                    try:
                        dt = datetime.fromisoformat(check_out.replace('Z', '+00:00'))
                        ws[f'D{row}'] = dt.strftime('%H:%M')
                    except:
                        ws[f'D{row}'] = check_out[:5] if len(check_out) > 5 else check_out
                else:
                    ws[f'D{row}'] = 'N/A'
                
                ws[f'E{row}'] = f"{att.get('working_hours', 0)}h"
                
                status = att.get('status', '')
                status_map = {
                    'present': 'Có mặt',
                    'late': 'Đi muộn',
                    'early_leave': 'Rời sớm',
                    'absent': 'Vắng mặt'
                }
                ws[f'F{row}'] = status_map.get(status, status)
                
                # Apply border
                for col in range(1, 7):
                    ws[f'{get_column_letter(col)}{row}'].border = border
                
                row += 1
        else:
            for day in report_data.get('daily_stats', []):
                ws[f'A{row}'] = day.get('date', '')
                ws[f'B{row}'] = day.get('present', 0)
                ws[f'C{row}'] = day.get('late', 0)
                ws[f'D{row}'] = day.get('absent', 0)
                if report_type == 'weekly':
                    ws[f'E{row}'] = f"{day.get('working_hours', 0)}h"
                else:
                    ws[f'E{row}'] = day.get('early_leave', 0)
                
                # Apply border
                for col in range(1, len(headers) + 1):
                    ws[f'{get_column_letter(col)}{row}'].border = border
                
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_num in range(1, row):
                cell_value = ws[f'{column_letter}{row_num}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

